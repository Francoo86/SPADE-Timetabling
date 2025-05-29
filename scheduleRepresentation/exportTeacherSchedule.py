import os
import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

SCRIPT_PATH = os.path.dirname(os.path.abspath(__file__))
PROJECT_PATH = os.path.join(SCRIPT_PATH, "..")

def create_teacher_schedule(teacher_data):
    """Creates a schedule DataFrame for a single teacher"""
    time_blocks = {
        1: "08:00 - 09:00",
        2: "09:15 - 10:15",
        3: "10:30 - 11:30",
        4: "11:45 - 12:45",
        5: "12:50 - 13:50",
        6: "13:55 - 14:55",
        7: "15:00 - 16:00",
        8: "16:15 - 17:15",
        9: "17:30 - 18:30",
    }
    
    days = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
    
    schedule_df = pd.DataFrame(
        index=time_blocks.values(),
        columns=days
    )
    
    # Fill schedule with teacher's subjects
    for subject in teacher_data['Asignaturas']:
        time_slot = time_blocks[subject['Bloque']]
        day = subject['Dia']
        content = (f"{subject['Nombre']}\n"
                  f"Sala: {subject['Sala']}\n"
                  )#f"Satisfacción: {subject['Satisfaccion']}/10")
        schedule_df.at[time_slot, day] = content
    
    return schedule_df.fillna('')

def apply_excel_styling(worksheet):
    """Applies styling to the Excel worksheet"""
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in worksheet.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:  # Headers
                cell.fill = header_fill
                cell.font = Font(bold=True)

def save_schedules(data, filename='teacher_schedules.xlsx'):
    """Saves each teacher's schedule to a separate worksheet in scheduleRepresentation folder"""
    output_dir = 'scheduleRepresentation'
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for teacher_data in data:
            teacher_name = teacher_data['Nombre']
            turno = teacher_data.get('Turno', 0)  # Get turno or default to 0
            schedule_df = create_teacher_schedule(teacher_data)
            
            sheet_name = f"{teacher_name}_{turno}"[:31]  # Include turno in sheet name
            schedule_df.to_excel(writer, sheet_name=sheet_name, index=True)
            
            worksheet = writer.sheets[sheet_name]
            apply_excel_styling(worksheet)
            
            # Add turno information to the worksheet
            worksheet.cell(row=1, column=1, value=f"Profesor: {teacher_name} (Turno: {turno})")
            
            for idx, col in enumerate(schedule_df.columns):
                worksheet.column_dimensions[chr(66 + idx)].width = 30
            worksheet.column_dimensions['A'].width = 15
            
            for row in range(1, len(schedule_df) + 2):
                worksheet.row_dimensions[row].height = 60

def main():
    try:
        with open(os.path.join(PROJECT_PATH, "agent_output/full/Horarios_asignados.json"), 'r', encoding="utf-8") as file:
            schedule_data = json.load(file)
            save_schedules(schedule_data)
            print(f"Schedules generated successfully for {len(schedule_data)} teachers")
            
            # Print summary
            for teacher in schedule_data:
                print(f"\nTeacher: {teacher['Nombre']}")
                print(f"Subjects assigned: {len(teacher['Asignaturas'])}/{teacher['Solicitudes']}")
    except Exception as e:
        print(f"Error processing schedules: {str(e)}")

if __name__ == '__main__':
    main()