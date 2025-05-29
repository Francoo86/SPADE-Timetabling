import os
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

SCRIPT_PATH = os.path.dirname(os.path.abspath(__file__))
PROJECT_PATH = os.path.join(SCRIPT_PATH, "..")

def create_room_schedule(room_data):
    """Creates a schedule DataFrame for a single room"""
    time_blocks = {
        1: "08:00 - 09:00",
        2: "09:15 - 10:15",
        3: "10:30 - 11:30",
        4: "11:45 - 12:45",
        5: "12:50 - 13:50",
        6: "13:55 - 14:55",
        7: "15:00 - 16:00",
        8: "16:15 - 17:15",
        9: "17:30 - 18:30",
    }
    days = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
    
    schedule_df = pd.DataFrame(
        index=time_blocks.values(),
        columns=days
    )
    
    # Fill schedule with room's subjects
    for subject in room_data.get('Asignaturas', []):
        time_slot = time_blocks.get(subject.get('Bloque'))
        if time_slot and subject.get('Dia') in days:
            content = (f"Asignatura: {subject.get('Nombre', 'Sin nombre')}\n"
                      # f"Satisfacción: {subject.get('Satisfaccion', 'N/A')}/10\n"
                      f"Capacidad: {subject.get('Capacidad', 0):.0%}")  # Mostrar capacidad como porcentaje
            schedule_df.at[time_slot, subject['Dia']] = content
    
    return schedule_df.fillna('')

def apply_excel_styling(worksheet):
    """Applies styling to the Excel worksheet"""
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFCCE5', end_color='FFCCE5', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in worksheet.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:  # Headers
                cell.fill = header_fill
                cell.font = Font(bold=True)

def save_room_schedules(data, filename='room_schedules.xlsx'):
    """Saves each room's schedule to a separate worksheet"""
    output_dir = 'scheduleRepresentation'
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    
    workbook = Workbook()
    first_sheet = True
    
    for room_data in data:
        room_code = room_data.get('Codigo', 'Sin código')
        campus = room_data.get('Campus', 'Sin campus')
        schedule_df = create_room_schedule(room_data)
        
        if first_sheet:
            sheet = workbook.active
            sheet.title = f"Sala {room_code}"[:31]
            first_sheet = False
        else:
            sheet = workbook.create_sheet(f"Sala {room_code}"[:31])
        
        # Write headers and info
        campus_cell = sheet.cell(row=1, column=1)
        campus_cell.value = f"Campus: {campus}"
        
        for col, day in enumerate(schedule_df.columns, start=2):
            sheet.cell(row=1, column=col, value=day)
        
        for row, time in enumerate(schedule_df.index, start=2):
            sheet.cell(row=row, column=1, value=time)
            for col, day in enumerate(schedule_df.columns, start=2):
                content = schedule_df.at[time, day]
                sheet.cell(row=row, column=col, value=content)
        
        apply_excel_styling(sheet)
        
        for col in range(1, len(schedule_df.columns) + 2):
            sheet.column_dimensions[chr(64 + col)].width = 30
        for row in range(1, len(schedule_df) + 2):
            sheet.row_dimensions[row].height = 80
    
    try:
        workbook.save(filepath)
        print(f"Excel file saved successfully at: {filepath}")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

def main():
    try:
        # Read the JSON file from the correct location
        json_path = os.path.join(PROJECT_PATH, "agent_output", "Horarios_salas.json")
        
        if not os.path.exists(json_path):
            print(f"Error: File not found at {json_path}")
            return
            
        with open(json_path, 'r', encoding="utf-8") as file:
            schedule_data = json.load(file)
            
            if not schedule_data:
                print("Error: No data found in JSON file")
                return
                
            save_room_schedules(schedule_data)
            print(f"Schedules generated successfully for {len(schedule_data)} rooms")
            
            # Print summary
            for room in schedule_data:
                print(f"\nRoom: {room.get('Codigo', 'Sin código')}")
                print(f"Capacity: {room.get('Capacidad', 'N/A')} students")
                print(f"Subjects assigned: {len(room.get('Asignaturas', []))}")
                
    except json.JSONDecodeError:
        print("Error: Invalid JSON format in the file")
    except Exception as e:
        print(f"Error processing schedules: {str(e)}")

if __name__ == '__main__':
    main()